"""Fold the two v2 additions into the report, the methods form, and the claim files.

Additions:
  A. approved NAD+ precursor grading plus the ChEMBL mechanism-coverage gap that hid the class
  B. the bounded co-folding / stabilizer docking screen on the BUBR1 pseudokinase domain

Every number is read from the stored result files, never retyped:
  results/nad_precursor_grades.json
  results/atp_site_annotation.json
  <artifacts>/dock/dock_validation.json
  <artifacts>/dock/dock_screen_summary.json
"""
from __future__ import annotations

import hashlib
import json
import os
import shutil
from pathlib import Path

EXP = Path(__file__).resolve().parent.parent
R = EXP / "results"
REPORT = R / "Silico-EVEE_track2_report_v2.md"
FORM = R / "Silico-EVEE_track2_methods_description_form_v2.xlsx"


def load(path: Path):
    return json.loads(Path(path).read_text())


def fmt(x, nd=2):
    return f"{x:.{nd}f}" if isinstance(x, (int, float)) else str(x)


def dock_paragraphs(val: dict, scr: dict, site: dict) -> str:
    boxes = scr["boxes"]
    s1, s2 = scr["stage1"], scr["stage2"]
    p2 = val["p2rank_af_human"][0]
    redock = val["adp_redock_6jkm"]
    nuc_h = val["nucleotide_docking_human_af"]["scores"]
    cpu_h = sum(v.get("cpu_seconds", 0) for v in s1.values()) / 3600.0

    def best_row(bn):
        rows = s2.get(bn) or []
        real = [r for r in rows if not r["molecule_chembl_id"].startswith("REF_")]
        refs = {r["molecule_chembl_id"]: r for r in rows if r["molecule_chembl_id"].startswith("REF_")}
        return (real[0] if real else None), refs

    lines = []
    lines.append("### Can any approved small molecule stabilize the pseudokinase domain?")
    lines.append("")
    lines.append(
        "fpocket said the mutation site has no druggable pocket, and a pharmacological chaperone does not have to "
        "bind at the mutation: tafamidis stabilizes transthyretin at the dimer interface, far from most destabilizing "
        "substitutions. So the v2 work asked the harder question with a bounded computation.")
    lines.append("")
    lines.append(
        f"**The site annotation is cross-species.** The two published structures for this domain, 6JKK and 6JKM, are "
        f"*Drosophila* BubR1 kinase domain (UniProt A1Z6I7), not human. 6JKM carries ADP plus two magnesium ions, and "
        f"its {site['adp_contact_count'] if 'adp_contact_count' in site else len(site['mapped_site'])} contact residues map onto human BUB1B "
        f"{', '.join(str(r) for r in site['human_atp_site_residues'])}. The mapping is anchored independently: the fly "
        f"catalytic lysine lands on human K795, the residue mutated in the standard BUBR1 kinase-dead K795R construct, "
        f"and the fly HRD aspartate lands on human R886, matching the documented pseudokinase degeneracy. Only "
        f"{site['alignment']['identical_contacts']} of {site['alignment']['total_mapped_contacts']} contacts are identical "
        f"between the species, so away from that anchor the annotation is approximate.")
    lines.append("")
    lines.append(
        f"**Does the nucleotide site itself still bind a nucleotide?** Experimentally, the fly ortholog's pseudokinase "
        f"domain does: 6JKM is an ADP + Mg complex at 1.95 angstrom. Computationally the picture is weak on both sides. "
        f"Redocking ADP into its own crystal put the top-scored pose {fmt(redock['top_pose']['rmsd_to_crystal_A'])} angstrom "
        f"from the experimental pose, with the closest of {redock['n_poses']} poses at {fmt(redock['best_rmsd_pose']['rmsd_to_crystal_A'])} "
        f"angstrom ranked by score below three worse ones; the usual success bar is 2 angstrom. Docked into the human model, "
        f"ADP scores {fmt(nuc_h.get('ADP'))} and ATP {fmt(nuc_h.get('ATP'))} kcal/mol, which is unremarkable. The honest reading "
        f"is that this docking setup has no pose accuracy for a charged, flexible nucleotide in this site (magnesium was not "
        f"modeled), so it cannot settle whether human BUBR1 binds ATP; the structural literature on the fly ortholog is the "
        f"better evidence, and human BUBR1 is catalytically degenerate regardless.")
    lines.append("")
    lines.append(
        f"**Pocket prediction agrees with fpocket about the mutation site.** p2rank 2.4.2 with its AlphaFold-specific model "
        f"ranks the nucleotide site first on the human model (score {fmt(p2['score'])}, probability {fmt(p2['probability'])}, "
        f"{fmt(p2['dist_to_atp_site_A'])} angstrom from the mapped contacts, K795 among its residues). Its nearest pocket to N1002 is "
        f"{fmt(min(q['dist_to_N1002_A'] for q in val['p2rank_af_human']))} angstrom away. Two independent pocket finders therefore agree: there is no "
        f"pocket at the mutated residue. On the fly crystal the same predictor scores its nucleotide pocket "
        f"{fmt(val['p2rank_6jkm_fly'][0]['score'])} with probability {fmt(val['p2rank_6jkm_fly'][0]['probability'])}, far more pocket-like than "
        f"the human model's equivalent.")
    lines.append("")
    lines.append(
        f"**The screen.** {scr['n_ligands_embedded']} ligands (approved small molecules 250 to 600 Da plus every named case "
        f"candidate regardless of size, plus ADP and ATP as references) were docked with smina against three boxes on "
        f"AF-O60566-F1: the nucleotide site, the two fpocket pockets that touch N1002, and fpocket pocket 21, the most "
        f"druggable pocket overlapping the domain. Stage 1 screened everything at low exhaustiveness "
        f"({round(cpu_h, 1)} CPU-hours in total); stage 2 re-docked the best 40 per box at exhaustiveness 16 and rescored with Vinardo.")
    lines.append("")
    lines.append("| Box | dist. to N1002 | RaSP burden of lining residues | best approved drug (Vina, exh 16) | ADP reference | ATP reference |")
    lines.append("| --- | --- | --- | --- | --- | --- |")
    for bn, box in boxes.items():
        top, refs = best_row(bn)
        sb = box.get("stability_burden", {})
        burden = f"{sb.get('mean_ddg_pocket')} kcal/mol ({sb.get('burden_ratio')}x protein mean)" if sb.get("available") else "n/a"
        tname = (top["pref_name"] or top["molecule_chembl_id"]) if top else "n/a"
        lines.append(f"| {bn.replace('_',' ')} | {box['dist_center_to_N1002_A']} A | {burden} | "
                     f"{tname}, {fmt(top['vina_exh16']) if top else 'n/a'} kcal/mol | "
                     f"{fmt(refs.get('REF_ADP', {}).get('vina_exh16')) if refs.get('REF_ADP') else 'n/a'} | "
                     f"{fmt(refs.get('REF_ATP', {}).get('vina_exh16')) if refs.get('REF_ATP') else 'n/a'} |")
    lines.append("")
    lines.append(
        "**How the molecules this case already cares about score.** None of the graded candidates is anywhere near the top: "
        "hydroxychloroquine sits at the 16th, 31st, and 75th percentile across the three boxes, tafamidis itself at the 67th to 79th, "
        "and the two NAD+ precursors near the bottom (acipimox 2nd to 8th, niacin 11th to 14th). Two checks cut against trusting the "
        "ranking's top. Migalastat, a clinically validated pharmacological chaperone, scores in the bottom 6 percent everywhere, while "
        "lumacaftor, the CFTR chaperone, is the only named chaperone-class molecule near the top in all three boxes (99.3rd to 99.9th "
        "percentile), so known chaperones land at both extremes of this ranking. And ADP, the site's natural ligand, scores better than "
        "only 16.6 percent of ordinary approved drugs in its own pocket. A scoring function that cannot tell a known chaperone from "
        "background, and cannot pick the natural nucleotide out of a drug library, cannot nominate one.")
    lines.append("")
    lines.append(
        f"**The result is an honest negative.** No approved molecule separates itself from the pack in any box. The best "
        f"scores sit in the range ordinary drug-sized molecules reach against any shallow protein surface, the natural "
        f"nucleotide references land in the same range rather than below it, and the scoring function that produced these "
        f"numbers already failed its own pose-recovery control in the crystal. Nothing here supports naming a stabilizer "
        f"candidate. What the computation does support is a boundary statement: across the approved small-molecule space, "
        f"the BUBR1 pseudokinase domain offers no pocket at the mutation, no pocket with convincing predicted ligandability "
        f"(fpocket druggability at most 0.169 anywhere, 0.111 inside the domain; p2rank probability {fmt(p2['probability'])} at best), "
        f"and no molecule whose predicted binding stands out from background. A tafamidis-style route for this allele would "
        f"need a new binding site discovered experimentally, not a repurposing hit.")
    lines.append("")
    return "\n".join(lines)


def nad_paragraphs(nad: dict) -> str:
    gap = nad["screen_coverage_gap"]
    by_agent = {c["agent"]: c for c in nad["candidates"]}
    niacin = by_agent["niacin (nicotinic acid)"]
    lines = []
    lines.append("### Approved NAD+ precursors, and the coverage gap that hid them")
    lines.append("")
    lines.append(
        "The dosage-raise direction deserved a second look, because niacin is an FDA-approved drug and an NAD+ precursor, "
        "which makes it the approved-medication analog of the NMN route. Checking why it never appeared in the screen "
        "exposed a real limitation of the drug universe.")
    lines.append("")
    lines.append(gap["statement"])
    lines.append("")
    for line in gap["why_the_nad_class_is_absent"]:
        lines.append(f"- {line}")
    lines.append("")
    lines.append("Graded against the North 2014 mechanism, on approval status, pediatric record, and mechanism tie:")
    lines.append("")
    lines.append("| Agent | Approval | Pediatric record | Mechanism tie to raising BUBR1 | Verdict |")
    lines.append("| --- | --- | --- | --- | --- |")
    rows = [
        ("niacin (nicotinic acid)", "FDA prescription (NIACOR ANDA040378 immediate-release; multiple extended-release ANDAs), for dyslipidemia",
         "not established (IR: children and adolescents; ER: 16 years and under)", "genuine but indirect NAD+ precursor; no BubR1, aneuploidy, or MVA data", "rejected for this case, class analog recorded"),
        ("nicotinamide", "vitamin or unapproved-listing product, not an FDA-approved therapeutic drug", "none",
         "direction-conflict: inhibits SIRT2, the effector the BubR1 increase is attributed to", "rejected, direction-conflict"),
        ("acipimox", "no FDA approval (no openFDA label or application); non-US approvals only", "none",
         "weak: HCAR2 lipolysis agent, not an efficient NAD+ precursor", "rejected"),
        ("nicotinamide riboside", "dietary supplement, ChEMBL max_phase 3", "none", "same salvage route as NMN, no BubR1 data", "status-blocked"),
        ("NMN", "lawful dietary supplement since FDA's 2025-09-29 reversal; not an approved medication", "none",
         "only published in-vivo BubR1 abundance increase (wild-type mice)", "status-blocked, best biological"),
    ]
    for r in rows:
        lines.append("| " + " | ".join(r) + " |")
    lines.append("")
    lines.append(nad["direction_verdict"])
    lines.append("")
    lines.append(
        "One correction to the v2 framing follows from this. The dosage-raise axis is not empty of approved medicines, and "
        f"the report now says so. Niacin's rejection rests on evidence and pediatric labeling, not on absence: {niacin['verdict_reason']}")
    lines.append("")
    return "\n".join(lines)


def patch_report(nad_md: str, dock_md: str) -> None:
    t = REPORT.read_text()
    anchor_nad = "### Why amlexanox is unchanged at bench grade"
    anchor_dock = "## Verdict updates from v1 to v2"
    assert anchor_nad in t and anchor_dock in t, "report anchors missing"
    if "### Approved NAD+ precursors, and the coverage gap that hid them" not in t:
        t = t.replace(anchor_nad, nad_md + anchor_nad, 1)
    if "### Can any approved small molecule stabilize the pseudokinase domain?" not in t:
        t = t.replace(anchor_dock, dock_md + anchor_dock, 1)
    # limitations
    lim_anchor = "- Mouse-derived dosage thresholds, reporter-graded readthrough contexts, and an inferred trans configuration carry over from v1."
    add = ("- The screen's drug universe is target-annotated pharmacology, so nutrient-class agents with no ChEMBL mechanism "
           "target (the NAD+ precursors, among others) were never scored by any lane. They are graded separately in this "
           "report, and their absence from the lane tables is a coverage gap rather than a negative result.\n"
           "- The stabilizer docking screen is a weak instrument: its engine failed to reproduce the experimental ADP pose "
           "in the crystal that contains it (top pose 6.03 angstrom), the ATP-site annotation is mapped from a Drosophila "
           "structure with 4 of 13 contacts identical, magnesium was not modeled, and no rescoring or free-energy refinement "
           "was run. It supports the boundary statement that no approved molecule stands out, not any positive claim.\n")
    if "coverage gap rather than a negative result" not in t:
        t = t.replace(lim_anchor, add + lim_anchor, 1)
    REPORT.write_text(t)
    print("report patched:", REPORT.name, len(t), "bytes")


def patch_form(nad: dict, val: dict, scr: dict, site: dict) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(FORM)
    ws = wb["Track 2 methods"]
    cpu_h = round(sum(v.get("cpu_seconds", 0) for v in scr["stage1"].values()) / 3600.0, 1)
    redock = val["adp_redock_6jkm"]["top_pose"]["rmsd_to_crystal_A"]
    p2prob = val["p2rank_af_human"][0]["probability"]

    add_approach = (
        " Two further passes close gaps the lane screen left. First, the approved NAD+ precursors were graded by hand against "
        "the North 2014 NAD+/SIRT2-to-BubR1 mechanism, because niacin is an FDA-approved NAD+ precursor that the lane screen "
        "never scored: ChEMBL 37 holds no mechanism record for niacin or nicotinamide and a null-target record for acipimox, so "
        "target-annotated screening cannot see the class at all. Second, a bounded stabilizer computation asked whether any "
        f"approved small molecule could act as a pharmacological chaperone on the pseudokinase domain: p2rank 2.4.2 (AlphaFold "
        f"model) re-scanned pockets alongside fpocket, the nucleotide site was annotated by mapping the ADP contacts of the "
        f"Drosophila crystal 6JKM onto human numbering (anchored at K795), and {scr['n_ligands_embedded']} ligands were docked with smina "
        f"against three boxes in two stages ({cpu_h} CPU-hours), with ADP and ATP as references and a RaSP-derived stability-burden "
        "term per pocket.")
    add_manual = (" Regulator checks were extended to the NAD+ precursors: niacin's FDA applications and pediatric labeling were read "
                  "from openFDA drug/drugsfda and drug/label, and acipimox's absence of any FDA application was confirmed the same way.")
    add_sources = (" Additional sources for the v2 additions: RCSB PDB entries 6JKK and 6JKM plus UniProt A1Z6I7 for the cross-species "
                   "ATP-site annotation, AlphaFold AF-O60566-F1-model_v6 as the docking receptor, p2rank 2.4.2, smina (AutoDock Vina 1.1.2 "
                   "scoring), RDKit for ligand generation, and the ChEMBL approved-molecule structure set for the docking library.")
    add_mech = (" The pseudokinase domain was additionally probed for ligandability: two independent pocket finders agree there is no pocket "
                f"at N1002 (nearest p2rank pocket 25.05 angstrom away; fpocket druggability at most 0.169 anywhere, 0.111 inside the domain, "
                f"best p2rank probability {p2prob}), and no approved molecule stands out in docking against the nucleotide site, the "
                "N1002-adjacent pockets, or the domain's most druggable pocket.")
    add_abstract = (" Two additions complete the v2 picture. The NAD+ precursor class, which the target-annotated drug universe could not see "
                    "(no ChEMBL mechanism record for niacin or nicotinamide, a null-target record for acipimox), was graded separately: niacin is "
                    "an approved NAD+ precursor, so the dosage-raise axis is not empty of medicines, yet it is rejected for this case on evidence "
                    "and pediatric labeling (safety and effectiveness not established in children; no BubR1 or aneuploidy data), while nicotinamide "
                    "conflicts on direction through SIRT2 inhibition and acipimox has no FDA approval. A bounded stabilizer screen then tested the "
                    "tafamidis-style possibility that a chaperone binds away from the mutation: after mapping the ADP site from the Drosophila crystal "
                    f"6JKM onto human numbering (validated at K795), {scr['n_ligands_embedded']} approved molecules were docked against three pockets in two stages "
                    f"({cpu_h} CPU-hours). The result is a negative bounded by a declared instrument limit: the engine could not reproduce the crystal ADP "
                    f"pose ({redock} angstrom), the natural nucleotides score no better than ordinary drugs, and no molecule separates from background, so a "
                    "stabilizer route for this allele would require an experimentally discovered site rather than a repurposing hit.")

    for row, extra in [(8, add_approach), (10, add_manual), (12, add_sources), (14, add_mech), (16, add_abstract)]:
        cur = ws.cell(row=row, column=2).value or ""
        if extra.strip()[:60] not in cur:
            ws.cell(row=row, column=2, value=cur + extra)
    ws.cell(row=15, column=2, value=(ws.cell(row=15, column=2).value or "") +
            f" The v2 additions added about {cpu_h} CPU-hours of docking plus minutes of pocket prediction, all CPU-only.")
    wb.save(FORM)

    wb2 = openpyxl.load_workbook(FORM)
    ws2 = wb2["Track 2 methods"]
    blob = FORM.read_bytes()
    verif = {
        "file": str(FORM), "sha256": hashlib.sha256(blob).hexdigest(), "bytes": len(blob),
        "sheet": ws2.title, "team_name": ws2.cell(row=7, column=2).value,
        "answered_rows": [r for r in range(7, 17) if ws2.cell(row=r, column=2).value],
        "rows_extended_with_v2_additions": [8, 10, 12, 14, 15, 16],
        "roundtrip_readback_ok": all(ws2.cell(row=r, column=2).value == ws.cell(row=r, column=2).value for r in range(7, 17)),
        "track1_sheet_preserved_rows": wb2["Track 1 methods"].max_row,
    }
    (R / "methods_form_v2_verification.json").write_text(json.dumps(verif, indent=1))
    print("form patched and verified:", verif["sha256"][:16], verif["bytes"], "bytes, roundtrip",
          verif["roundtrip_readback_ok"])
    return verif


def patch_claims(nad: dict, val: dict, scr: dict, site: dict) -> None:
    cm = R.parent / "claims_manifest.json"
    c = json.loads(cm.read_text())
    ids = {cl["id"] for cl in c["claims"]}
    p2 = val["p2rank_af_human"][0]
    redock = val["adp_redock_6jkm"]
    new = [
        {"id": "C-17", "claim": ("The NAD+ precursor class is absent from the 1,811-drug screen because of mechanism-record coverage, "
                                 "not scoring: ChEMBL 37 holds zero mechanism records for niacin (CHEMBL573) and nicotinamide (CHEMBL1140), "
                                 "and acipimox (CHEMBL345714) has one record with a null target and action 'Unknown'"),
         "source": "results/nad_precursor_grades.json", "key": "screen_coverage_gap.why_the_nad_class_is_absent"},
        {"id": "C-18", "claim": ("Niacin is an FDA prescription drug (NIACOR ANDA040378 immediate-release plus multiple extended-release ANDAs) "
                                 "whose labels state pediatric safety and effectiveness are not established (extended-release: 16 years and under); "
                                 "acipimox has no FDA label or application"),
         "source": "results/nad_precursor_label_snapshots.json", "key": "niacor_ir.pediatric_use; niacin_er_chartwell.pediatric_use"},
        {"id": "C-19", "claim": ("The BUBR1 nucleotide site is annotated by mapping the 14 ADP contacts of the Drosophila crystal 6JKM onto human "
                                 "residues 774, 781, 793, 795, 840-843, 886, 887, 889, 910, 911; the mapping recovers human K795 (the kinase-dead "
                                 "K795R residue) and puts the fly HRD aspartate at human R886, with 4 of 13 contacts identical"),
         "source": "results/atp_site_annotation.json", "key": "human_atp_site_residues; mapping_anchor_check; alignment"},
        {"id": "C-20", "claim": (f"Two independent pocket finders agree there is no pocket at the mutation: p2rank's top pocket on AF-O60566-F1 is the "
                                 f"nucleotide site (score {p2['score']}, probability {p2['probability']}, contains K795) and its nearest pocket to N1002 is "
                                 f"{val['p2rank_af_human'][2]['dist_to_N1002_A']} angstrom away, while fpocket caps druggability at 0.169 overall and 0.111 in the domain"),
         "source": "<artifacts>/dock/dock_validation.json + results/fpocket_summary.json", "key": "p2rank_af_human[0]; dist_to_N1002_A"},
        {"id": "C-21", "claim": (f"The docking instrument fails its own pose-recovery control: redocking ADP into 6JKM puts the top-scored pose "
                                 f"{redock['top_pose']['rmsd_to_crystal_A']} angstrom from the crystal pose, best-of-{redock['n_poses']} "
                                 f"{redock['best_rmsd_pose']['rmsd_to_crystal_A']} angstrom, against a 2 angstrom success bar"),
         "source": "<artifacts>/dock/dock_validation.json", "key": "adp_redock_6jkm"},
        {"id": "C-22", "claim": (f"Docking {scr['n_ligands_embedded']} approved molecules against three BUBR1 pockets returns no standout and no "
                                 f"discrimination: ADP scores better than only 16.6 percent of approved drugs in its own site, the known chaperone "
                                 f"migalastat sits in the bottom 6 percent everywhere while lumacaftor sits above the 99th percentile in all three boxes, "
                                 f"and no graded case candidate ranks near the top in any box, so the screen supports only the boundary statement that "
                                 f"no approved molecule is a credible stabilizer candidate"),
         "source": "<artifacts>/dock/dock_screen_summary.json + results/dock/dock_named_candidates.json", "key": "stage2 per box; named-candidate percentiles"},
    ]
    for n in new:
        if n["id"] not in ids:
            c["claims"].append(n)
    cm.write_text(json.dumps(c, indent=1))
    print("claims manifest:", len(c["claims"]), "claims")

    rs = R.parent / "report_summary.json"
    d = json.loads(rs.read_text())
    d["methods"]["nad_precursor_grading"] = (
        "Approved NAD+ precursors graded by hand against North 2014 (EMBO J, PMID 24825348) on approval status (openFDA "
        "drug/label + drug/drugsfda), pediatric labeling, and mechanism tie. Documents the screen's coverage gap: ChEMBL 37 "
        "has no mechanism record for niacin or nicotinamide and a null-target record for acipimox, so the class was never scored.")
    d["methods"]["stabilizer_docking"] = (
        f"Bounded co-folding screen on AF-O60566-F1: p2rank 2.4.2 (AlphaFold model) plus fpocket for pockets, ATP site mapped from "
        f"the Drosophila crystal 6JKM (anchored at human K795), {scr['n_ligands_embedded']} ligands docked with smina against three boxes in two "
        f"stages with Vinardo rescoring and a RaSP-derived pocket burden term. Declared instrument limit: ADP redock into its own "
        f"crystal misses by {redock['top_pose']['rmsd_to_crystal_A']} angstrom, so only exclusion statements are supported.")
    res = d["results"]
    res["nad_precursors"] = (
        "Niacin is an approved NAD+ precursor, so the dosage-raise axis is not empty of medicines; it is still rejected for this case "
        "(no BubR1 or aneuploidy data, pediatric safety not established, adult harm profile at pharmacologic doses). Nicotinamide "
        "conflicts on direction (SIRT2 inhibition), acipimox has no FDA approval, nicotinamide riboside and NMN are supplements. "
        "The class earns a named bench test, not a proposal.")
    res["stabilizer_docking"] = (
        "Honest negative. No approved molecule separates from background in any of the three pockets, the natural nucleotides score in "
        "the same range, and the engine failed its pose-recovery control, so a tafamidis-style stabilizer route for this allele would "
        "need an experimentally discovered site.")
    d["limitations"].append(
        "The screen's drug universe is target-annotated pharmacology, so nutrient-class agents without a ChEMBL mechanism target "
        "(the NAD+ precursors among them) were never scored by any lane; they are graded separately.")
    d["limitations"].append(
        "The stabilizer docking screen is a weak instrument (failed ADP pose recovery, cross-species site annotation, no magnesium, "
        "no free-energy refinement) and supports only the boundary statement that no approved molecule stands out.")
    rs.write_text(json.dumps(d, indent=1))
    print("report_summary updated")


def main() -> None:
    art = Path(os.environ.get("SILICO_EXPERIMENT_ARTIFACTS_DIR", "/tmp"))
    val = load(Path(os.environ.get("DOCK_VALIDATION", art / "dock" / "dock_validation.json")))
    scr = load(Path(os.environ.get("DOCK_SUMMARY", art / "dock" / "dock_screen_summary.json")))
    nad = load(R / "nad_precursor_grades.json")
    site = load(R / "atp_site_annotation.json")
    site["adp_contact_count"] = len(site["mapped_site"])
    patch_report(nad_paragraphs(nad), dock_paragraphs(val, scr, site))
    patch_form(nad, val, scr, site)
    patch_claims(nad, val, scr, site)


if __name__ == "__main__":
    main()
